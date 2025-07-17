from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import time
import re

FILE_PATH = "test_results.xlsx"

if os.path.exists(FILE_PATH):
    os.remove(FILE_PATH)
    print("🗑️ Deleted old test_results.xlsx")

account_balances = {
    "800000": 1000.00,
    "800001": 1000.00,
}

def setup_workbook():
    wb = Workbook()
    ws_balance = wb.active
    ws_balance.title = "AccountBalances"
    ws_balance.append(["Function", "Account", "Status"])

    ws_txn = wb.create_sheet("TransactionHistory")
    ws_txn.append(["Function", "Transaction ID", "Account", "Date", "Action", "Amount", "Status"])

    ws_fx = wb.create_sheet("CurrencyExchange")
    ws_fx.append(["Function", "From Currency", "To Currency", "Amount", "Result", "Status"])

    ws_transfer = wb.create_sheet("Transfers")
    ws_transfer.append(["Function", "From Account", "To Account", "Amount", "Status"])

    ws_topup = wb.create_sheet("TopUps")
    ws_topup.append(["Function","Account", "Phone Number", "Amount", "Network", "Status"])

    ws_pin = wb.create_sheet("PinCodes")
    ws_pin.append(["Function", "Account", "PIN Code", "Amount", "Status"])

    return wb, ws_balance, ws_txn, ws_fx, ws_transfer, ws_topup, ws_pin

def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

def update_account_balances_sheet(ws_balance, balances, label="RemainingBalance"):
    ws_balance.append([])
    ws_balance.append([f"{label}", "", ""])
    for acc, bal in balances.items():
        ws_balance.append([label, acc, f"${round(bal, 2)}"])

def check_transactions(driver, ws_txn):
    wait = WebDriverWait(driver, 15)
    function_name = "CheckTransactions"
    try:
        wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "View Recent Transactions"))).click()
        rows = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="_ctl0__ctl0_Content_Main_MyTransactions"]/tbody/tr[not(@style)]')))
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 5:
                txn_id = cells[0].text.strip()
                date = cells[1].text.strip()
                account = cells[2].text.strip()
                action = cells[3].text.strip()
                amount = cells[4].text.strip()
                ws_txn.append([function_name, txn_id, account, date, action, amount, "Success"])
    except:
        ws_txn.append([function_name, "", "", "", "", "", "Error"])

def exchange_currency(ws_fx):
    function_name = "CurrencyExchange"
    from_currency = "USD"
    to_currency = "KHR"
    amount = 200
    exchange_rate = 4095
    result = round(amount * exchange_rate, 2)
    try:
        ws_fx.append([function_name, from_currency, to_currency, amount, result, "Success"])
    except:
        ws_fx.append([function_name, from_currency, to_currency, amount, "", "Error"])

def transfer_funds(driver, ws_transfer, ws_txn):
    wait = WebDriverWait(driver, 15)
    function_name = "TransferFunds"
    txn_function = "CheckTransactions"
    transfers = [
        {"from": "800001", "to": "800000", "amount": "100.00"},
        {"from": "800000", "to": "800001", "amount": "250.00"},
        {"from": "800001", "to": "800000", "amount": "-10.00"},
        {"from": "800000", "to": "800001", "amount": "0.00"}
    ]
    for transfer in transfers:
        from_acc, to_acc, amount = transfer.values()
        amt_float = float(amount)
        try:
            wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Transfer Funds"))).click()
            Select(driver.find_element(By.ID, "fromAccount")).select_by_value(from_acc)
            Select(driver.find_element(By.ID, "toAccount")).select_by_value(to_acc)
            driver.find_element(By.ID, "transferAmount").clear()
            driver.find_element(By.ID, "transferAmount").send_keys(amount)
            driver.find_element(By.ID, "transfer").click()
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            txn_id = f"TX{int(time.time() * 1000)}"
            if "was successfully transferred" in driver.page_source:
                ws_transfer.append([function_name, from_acc, to_acc, amount, "Success"])
                account_balances[from_acc] -= amt_float
                account_balances[to_acc] += amt_float
                ws_txn.append([txn_function, txn_id, from_acc, timestamp, f"Transfer to {to_acc}", f"-{amount}", "Success"])
                ws_txn.append([txn_function, txn_id, to_acc, timestamp, f"Received from {from_acc}", f"+{amount}", "Success"])
            else:
                ws_transfer.append([function_name, from_acc, to_acc, amount, "Failed"])
        except:
            ws_transfer.append([function_name, from_acc, to_acc, amount, "Error"])

def top_up_mobile(driver, ws_topup, ws_txn):
    wait = WebDriverWait(driver, 15)
    function_name = "TopUp"
    txn_function = "CheckTransactions"
    topups = [
        {"phone": "010123456", "amount": 2.0, "network": "Cellcard"},
        {"phone": "097654321", "amount": 5.0, "network": "Metfone"},
        {"phone": "088888888", "amount": 10.0, "network": "Smart"},
    ]
    for i, topup in enumerate(topups):
        account_used = "800000" if i % 2 == 0 else "800001"
        phone, amount, network = topup.values()
        if account_balances[account_used] < amount:
            ws_topup.append([function_name, account_used,phone, amount, network, "Insufficient Funds"])
            continue
        try:
            driver.get("file:///D:/DaneChum/test/topup.html")
            wait.until(EC.presence_of_element_located((By.ID, "phoneNumber"))).send_keys(phone)
            Select(driver.find_element(By.ID, "network")).select_by_visible_text(network)
            driver.find_element(By.ID, "amount").send_keys(str(int(amount)))
            driver.find_element(By.ID, "submitBtn").click()
            if "Top-up successful" in driver.page_source:
                account_balances[account_used] -= amount
                ws_topup.append([function_name, account_used,phone, amount, network, "Success"])
                txn_id = f"TP{int(time.time() * 1000)}"
                date_str = time.strftime("%Y-%m-%d %H:%M:%S")
                ws_txn.append([txn_function, txn_id, account_used, date_str, f"Top-up {phone}", f"-{amount}", "Success"])
            else:
                ws_topup.append([function_name,account_used, phone, amount, network, "Failed"])
        except:
            ws_topup.append([function_name,account_used,phone, amount, network, "Error"])

def buy_pincode(driver, ws_pin, ws_txn):
    wait = WebDriverWait(driver, 15)
    function_name = "BuyPIN"
    txn_function = "CheckTransactions"
    purchases = [2.0, 5.0, 10.0]
    for i, amount in enumerate(purchases):
        account_used = "800000" if i % 2 == 0 else "800001"
        if account_balances[account_used] < amount:
            ws_pin.append([function_name, account_used, "", amount, "Insufficient Funds"])
            continue
        try:
            driver.get("file:///D:/DaneChum/test/pincode.html")
            wait.until(EC.presence_of_element_located((By.ID, "amount"))).send_keys(str(int(amount)))
            driver.find_element(By.ID, "phoneNumber").send_keys("012345678")
            driver.find_element(By.ID, "buyBtn").click()
            result_element = wait.until(EC.presence_of_element_located((By.ID, "result")))
            result_text = result_element.text.strip()
            if "PIN purchase successful" in result_text:
                pin_code = re.search(r"PIN Code:\s*(\w+)", result_text)
                pin_code = pin_code.group(1) if pin_code else "N/A"
                account_balances[account_used] -= amount
                ws_pin.append([function_name, account_used, pin_code, amount, "Success"])
                txn_id = f"PN{int(time.time() * 1000)}"
                date_str = time.strftime("%Y-%m-%d %H:%M:%S")
                ws_txn.append([txn_function, txn_id, account_used, date_str, "PIN Code Purchase", f"-{amount}", "Success"])
            else:
                ws_pin.append([function_name, account_used, "", amount, "Failed"])
        except:
            ws_pin.append([function_name, account_used, "", amount, "Error"])

def run_script():
    username = "admin"
    password = "admin"
    wb, ws_balance, ws_txn, ws_fx, ws_transfer, ws_topup, ws_pin = setup_workbook()
    driver = webdriver.Chrome()
    driver.maximize_window()
    wait = WebDriverWait(driver, 15)
    try:
        driver.get("https://demo.testfire.net/")
        driver.find_element(By.ID, "LoginLink").click()
        wait.until(EC.presence_of_element_located((By.NAME, "uid")))
        driver.find_element(By.NAME, "uid").send_keys(username)
        driver.find_element(By.NAME, "passw").send_keys(password)
        driver.find_element(By.NAME, "btnSubmit").click()
        if "Hello Admin User" not in driver.page_source:
            ws_balance.append(["Login", "", "Failed"])
            return
        update_account_balances_sheet(ws_balance, account_balances, label="InitialBalance")
        check_transactions(driver, ws_txn)
        exchange_currency(ws_fx)
        transfer_funds(driver, ws_transfer, ws_txn)
        update_account_balances_sheet(ws_balance, account_balances, label="PostTransferBalance")
        top_up_mobile(driver, ws_topup, ws_txn)
        update_account_balances_sheet(ws_balance, account_balances, label="PostTopUpBalance")
        buy_pincode(driver, ws_pin, ws_txn)
        update_account_balances_sheet(ws_balance, account_balances, label="PostPinBalance")
    finally:
        for sheet in [ws_balance, ws_txn, ws_fx, ws_transfer, ws_topup, ws_pin]:
            auto_fit_columns(sheet)
        wb.save(FILE_PATH)
        driver.quit()

run_script()
