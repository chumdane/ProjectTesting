from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time

# Load Excel file
wb = openpyxl.load_workbook("login_data.xlsx")
sheet = wb.active

# Setup Selenium WebDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Login page URL (replace with your actual target)
login_url = "https://example.com/login"

# Loop through each user
for row in range(2, sheet.max_row + 1):
    username = sheet.cell(row=row, column=1).value
    password = sheet.cell(row=row, column=2).value

    # Go to login page
    driver.get(login_url)
    time.sleep(2)

    # Fill in login fields (replace ID values with actual ones)
    driver.find_element(By.ID, "username").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()
    time.sleep(2)

    # Validate login success
    if "dashboard" in driver.current_url or "Welcome" in driver.page_source:
        result = "Success"
    else:
        result = "Failed"

    # Save result to Excel
    sheet.cell(row=row, column=3).value = result
    print(f"{username} => {result}")

# Save updated Excel file
wb.save("login_data_result.xlsx")

# Close browser
driver.quit()
print("✅ Test complete. Results saved in 'login_data_result.xlsx'")
